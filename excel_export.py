"""
Builds a downloadable .xlsx workbook for a computed roll-out plan and its
(optional) session plan, styled to match the real SETA roll-out plan
template — navy header row, bold merged Module column, wide wrapped
Component column, thin borders throughout.

Two distinct sheets in one workbook, per request: "Computed Schedule"
(named after the cohort) and "Session Plan" — kept structurally separate
rather than stacked in one sheet.
"""
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY = "FF002060"
WHITE = "FFFFFFFF"
MUTED = "FF787774"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

MODULE_FONT = Font(name="Arial", size=10, bold=True)
MODULE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Arial", size=14, bold=True)
META_FONT = Font(name="Arial", size=9, color=MUTED)
FOOTNOTE_FONT = Font(name="Arial", size=8, italic=True, color=MUTED)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _sheet_safe_name(name, fallback):
    """Excel sheet names: max 31 chars, and can't contain : \\ / ? * [ ]"""
    name = (name or "").strip()
    if not name:
        name = fallback
    name = re.sub(r'[:\\/?*\[\]]', '-', name)
    return name[:31] or fallback


def _split_module_component(name):
    """
    Splits a row name like "Module X: Unit Standard Y" (the format the
    xlsx-upload importer produces) into (module, component). Rows without
    that separator (e.g. manually typed) are treated as their own
    single-row module — module=name, component="".
    """
    name = name or ""
    if ": " in name:
        module, component = name.split(": ", 1)
        return module.strip(), component.strip()
    return name.strip(), ""


def _style(cell, font=BODY_FONT, align=None, border=BORDER, fill=None):
    cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if fill:
        cell.fill = fill


def _build_schedule_sheet(wb, cohort_name, induction_date, start_date, exam_date, modules):
    ws = wb.create_sheet(_sheet_safe_name(cohort_name, "Computed Schedule"))

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    row = 1
    ws.cell(row=row, column=1, value=cohort_name or "Roll-out Plan").font = TITLE_FONT
    row += 1

    meta_bits = []
    if induction_date:
        meta_bits.append(f"Induction: {induction_date}")
    if start_date:
        meta_bits.append(f"Programme Start: {start_date}")
    if exam_date:
        meta_bits.append(f"Exam Deadline: {exam_date}")
    if meta_bits:
        ws.cell(row=row, column=1, value="   |   ".join(meta_bits)).font = META_FONT
        row += 1
    row += 1  # spacer

    header_row = row
    headers = ["Module", "Unit Standard / Component", "Credits", "Notional Hours",
               "% of Programme", "Allocated Days", "Start Date", "End Date"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        _style(c, font=HEADER_FONT, align=HEADER_ALIGN, fill=HEADER_FILL)
    ws.row_dimensions[header_row].height = 32

    r = header_row + 1
    prev_module = None
    merge_start = None

    def close_merge(end_row):
        nonlocal merge_start
        if merge_start is not None and end_row > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=end_row, end_column=1)
        merge_start = None

    for m in modules:
        module_name, component = _split_module_component(m.get("name", ""))
        if module_name != prev_module:
            close_merge(r - 1)
            merge_start = r
            prev_module = module_name

        _style(ws.cell(row=r, column=1, value=module_name), font=MODULE_FONT, align=MODULE_ALIGN)
        _style(ws.cell(row=r, column=2, value=component or module_name), align=WRAP_LEFT)
        _style(ws.cell(row=r, column=3, value=m.get("credits")), align=CENTER)
        _style(ws.cell(row=r, column=4, value=m.get("notionalHours")), align=CENTER)

        pct = m.get("pct")
        pct_val = f"{pct:.1f}%" if isinstance(pct, (int, float)) else pct
        _style(ws.cell(row=r, column=5, value=pct_val), align=CENTER)

        _style(ws.cell(row=r, column=6, value=m.get("days")), align=CENTER)
        _style(ws.cell(row=r, column=7, value=m.get("start") or "\u2014"), align=CENTER)
        _style(ws.cell(row=r, column=8, value=m.get("end") or "\u2014"), align=CENTER)

        ws.row_dimensions[r].height = 30
        r += 1

    close_merge(r - 1)

    footer_row = r + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    ws.cell(row=footer_row, column=1,
            value=("Notional Hours is shown for SETA/assessor reference only — the schedule above "
                   "is driven by proportional business-day allocation by credit share, not notional hours.")
            ).font = FOOTNOTE_FONT

    return ws


def _build_sessions_sheet(wb, sessions):
    ws = wb.create_sheet("Session Plan")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16

    row = 1
    any_written = False

    for mod in sessions or []:
        rows_data = mod.get("rows") or []
        if not rows_data:
            continue
        any_written = True

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _style(
            ws.cell(row=row, column=1, value=mod.get("name", "")),
            font=Font(name="Arial", size=11, bold=True, color=WHITE),
            align=Alignment(horizontal="left", vertical="center"),
            fill=HEADER_FILL,
        )
        ws.row_dimensions[row].height = 22
        row += 1

        meta = f"{mod.get('start', '')} \u2013 {mod.get('end', '')}    |    {len(rows_data)} session(s)"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=meta).font = META_FONT
        row += 1

        header_row = row
        for col_idx, h in enumerate(["#", "Session Date", "Day"], start=1):
            _style(ws.cell(row=header_row, column=col_idx, value=h), font=HEADER_FONT, align=HEADER_ALIGN, fill=HEADER_FILL)
        row += 1

        for sess in rows_data:
            _style(ws.cell(row=row, column=1, value=sess.get("n")), align=CENTER)
            _style(ws.cell(row=row, column=2, value=sess.get("date")), align=CENTER)
            _style(ws.cell(row=row, column=3, value=sess.get("day")), align=CENTER)
            row += 1

        row += 2  # spacer between module blocks

    if not any_written:
        ws.cell(row=1, column=1, value="No sessions were generated before exporting — go back to the Session Planner, generate dates, then export again.").font = META_FONT

    return ws


def build_workbook(cohort_name, induction_date, start_date, exam_date, modules, sessions):
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    _build_schedule_sheet(wb, cohort_name, induction_date, start_date, exam_date, modules)
    _build_sessions_sheet(wb, sessions)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()